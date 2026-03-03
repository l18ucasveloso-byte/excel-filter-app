import streamlit as st
import pandas as pd
import io
import re
from openpyxl import load_workbook

# ==========================
# Configuração básica
# ==========================

st.set_page_config(
    page_title="Processador de Planilhas Excel",
    page_icon="📊",
    layout="wide"
)

MAX_FILES = 10
MAX_SHEETS_PER_FILE = 20
MAX_ROWS_PER_SHEET = 100_000
MAX_FILE_SIZE_MB = 50


# ==========================
# Funções auxiliares
# ==========================

def parse_keywords(raw_text: str):
    if not raw_text:
        return []
    return [k.strip() for k in raw_text.split(",") if k.strip()]


def build_primary_mask(df: pd.DataFrame, keywords):
    if df.empty or not keywords:
        return pd.Series([False] * len(df), index=df.index)

    pattern = "|".join(re.escape(k) for k in keywords)
    combined = df.astype(str).agg(" ".join, axis=1)
    return combined.str.contains(pattern, case=False, na=False)


def classify_rows(df: pd.DataFrame):
    if df.empty:
        empty_df = pd.DataFrame(columns=list(df.columns))
        return empty_df, empty_df, empty_df

    combined = df.astype(str).agg(" ".join, axis=1).str.lower()

    mask_cert = combined.str.contains("certificado", na=False)
    mask_log = combined.str.contains("logbook", na=False)

    df_certificados = df[mask_cert].copy()
    df_logbook = df[mask_log].copy()
    df_resto = df[~(mask_cert | mask_log)].copy()

    return df_certificados, df_logbook, df_resto


def df_to_excel_bytes(df: pd.DataFrame, sheet_name: str):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
    output.seek(0)
    return output


def get_sheet_names(uploaded_file):
    uploaded_file.seek(0)
    wb = load_workbook(uploaded_file, read_only=True, data_only=True)
    sheet_names = wb.sheetnames
    wb.close()
    uploaded_file.seek(0)
    return sheet_names


def validate_files(uploaded_files):
    if len(uploaded_files) > MAX_FILES:
        st.error(f"Máximo permitido: {MAX_FILES} arquivos.")
        return False

    valid_files = []
    for f in uploaded_files:
        size_mb = f.size / (1024 * 1024)
        if size_mb > MAX_FILE_SIZE_MB:
            st.error(
                f"Arquivo '{f.name}' tem {size_mb:.1f} MB. "
                f"Máximo permitido: {MAX_FILE_SIZE_MB} MB."
            )
            continue
        valid_files.append(f)

    if not valid_files:
        st.error("Nenhum arquivo válido para processamento.")
        return False

    return True


# ==========================
# Processamento principal (VERSÃO FINAL)
# ==========================

def process_files(uploaded_files, keywords):
    all_certificados = []
    all_logbook = []
    all_resto = []

    total_files = len(uploaded_files)
    progress_bar = st.progress(0.0)
    status_text = st.empty()

    for i, uploaded_file in enumerate(uploaded_files, start=1):
        try:
            status_text.text(f"📂 Arquivo {i}/{total_files}: {uploaded_file.name}")
            
            sheet_names = get_sheet_names(uploaded_file)

            if not sheet_names:
                st.warning(f"⚠️ '{uploaded_file.name}' não possui abas.")
                progress_bar.progress(i / total_files)
                continue

            if len(sheet_names) > MAX_SHEETS_PER_FILE:
                st.warning(
                    f"⚠️ '{uploaded_file.name}': {len(sheet_names)} abas → "
                    f"processando apenas {MAX_SHEETS_PER_FILE}"
                )
                sheet_names = sheet_names[:MAX_SHEETS_PER_FILE]

            for j, sheet_name in enumerate(sheet_names, start=1):
                try:
                    st.info(
                        f"📋 '{uploaded_file.name}' → aba {j}/{len(sheet_names)}: '{sheet_name}'"
                    )

                    uploaded_file.seek(0)
                    df = pd.read_excel(
                        uploaded_file,
                        sheet_name=sheet_name,
                        engine="openpyxl",
                        dtype=str
                    )

                    if df.empty:
                        continue

                    # Limite de linhas com aviso
                    if len(df) > MAX_ROWS_PER_SHEET:
                        st.warning(
                            f"⚠️ Aba '{sheet_name}': {len(df):,} linhas → "
                            f"limitando a {MAX_ROWS_PER_SHEET:,}"
                        )
                        df = df.iloc[:MAX_ROWS_PER_SHEET].copy()

                    # Rastreabilidade
                    df = df.copy()
                    df["Numero_Linha_Original"] = df.index + 2
                    df["Nome_Arquivo_Origem"] = uploaded_file.name

                    # Busca principal
                    mask_primary = build_primary_mask(df, keywords)
                    df_filtered = df[mask_primary]

                    if df_filtered.empty:
                        del df
                        continue

                    # Classificação
                    df_cert, df_log, df_rest = classify_rows(df_filtered)

                    if not df_cert.empty:
                        all_certificados.append(df_cert)
                    if not df_log.empty:
                        all_logbook.append(df_log)
                    if not df_rest.empty:
                        all_resto.append(df_rest)

                    # 🧹 Libera memória
                    del df, df_filtered, df_cert, df_log, df_rest

                except Exception as e_sheet:
                    st.error(f"❌ Aba '{sheet_name}': {e_sheet}")
                    continue

        except Exception as e_file:
            st.error(f"❌ Arquivo '{uploaded_file.name}': {e_file}")

        progress_bar.progress(i / total_files)

    status_text.text("✅ Processamento concluído!")

    # Concatenação segura
    df_certificados_final = (
        pd.concat(all_certificados, ignore_index=True)
        if all_certificados else pd.DataFrame()
    )
    df_logbook_final = (
        pd.concat(all_logbook, ignore_index=True)
        if all_logbook else pd.DataFrame()
    )
    df_resto_final = (
        pd.concat(all_resto, ignore_index=True)
        if all_resto else pd.DataFrame()
    )

    return df_certificados_final, df_logbook_final, df_resto_final


# ==========================
# Interface Streamlit
# ==========================

def main():
    st.title("📊 Processador de Planilhas Excel")

    st.markdown("""
**Limites de uso:**
- 🔢 10 arquivos máximo
- 📁 50 MB por arquivo  
- 📋 20 abas por arquivo
- 📈 100.000 linhas por aba
    """)

    st.sidebar.header("⚙️ Configurações")

    uploaded_files = st.sidebar.file_uploader(
        "Envie arquivos .xlsx",
        type=["xlsx"],
        accept_multiple_files=True
    )

    raw_keywords = st.sidebar.text_area(
        "🔍 Palavras-chave (separadas por vírgula)",
        placeholder="Ex: CASE12, CAT998, TESTE123"
    )

    process_button = st.sidebar.button("🚀 Processar", type="primary")

    if process_button:
        if not uploaded_files:
            st.error("📤 Envie pelo menos um arquivo.")
            return

        if not validate_files(uploaded_files):
            return

        keywords = parse_keywords(raw_keywords)
        if not keywords:
            st.error("🔍 Informe pelo menos uma palavra-chave.")
            return

        with st.spinner("🔄 Processando arquivos..."):
            df_certificados, df_logbook, df_resto = process_files(
                uploaded_files, keywords
            )

        st.subheader("📋 Resultados")

        total_rows = len(df_certificados) + len(df_logbook) + len(df_resto)
        if total_rows == 0:
            st.warning("❌ Nenhum resultado encontrado para as palavras-chave informadas.")
            return

        st.success(f"""
✅ **{total_rows:,} linhas encontradas:**

| Categoria | Quantidade |
|-----------|------------|
| 📄 Certificados | {len(df_certificados):,} |
| 📘 Logbook | {len(df_logbook):,} |
| 📂 Resto | {len(df_resto):,} |
        """)

        col1, col2, col3 = st.columns(3)

        with col1:
            st.markdown("### 📄 **Certificados**")
            if not df_certificados.empty:
                st.dataframe(df_certificados.head(20), use_container_width=True)
                st.download_button(
                    "⬇️ Baixar Certificados.xlsx",
                    data=df_to_excel_bytes(df_certificados, "Certificados"),
                    file_name="Certificados.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            else:
                st.info("Nenhum certificado encontrado.")

        with col2:
            st.markdown("### 📘 **Logbook**")
            if not df_logbook.empty:
                st.dataframe(df_logbook.head(20), use_container_width=True)
                st.download_button(
                    "⬇️ Baixar Logbook.xlsx",
                    data=df_to_excel_bytes(df_logbook, "Logbook"),
                    file_name="Logbook.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            else:
                st.info("Nenhum logbook encontrado.")

        with col3:
            st.markdown("### 📂 **Resto**")
            if not df_resto.empty:
                st.dataframe(df_resto.head(20), use_container_width=True)
                st.download_button(
                    "⬇️ Baixar Resto.xlsx",
                    data=df_to_excel_bytes(df_resto, "Resto"),
                    file_name="Resto.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            else:
                st.info("Nenhum resultado restante.")


if __name__ == "__main__":
    main()
